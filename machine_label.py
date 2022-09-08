import openpyxl
from datasets import load_dataset
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

GREEN = '7FFF00'
RED = 'EE4B2B'
DATA_PATH = "./data"
LOG_FILE = DATA_PATH + "/log.txt"


def write_label(answers, file_index):
    file_name = "qa_sample_" + str(file_index).zfill(4)
    full_path = DATA_PATH + "/" + file_name + ".xlsx"

    workbook = openpyxl.load_workbook(full_path)
    worksheet = workbook["Sheet1"]
    num_label, num_cor = 0, 0
    num_qa = len(answers)
    for i, answer in enumerate(answers):
        machine_cell = worksheet.cell(row=i + 2, column=4)
        machine_cell.value = answer["answer"]
        machine_cell.alignment = Alignment(wrap_text=True, vertical='top')
        human_answer = worksheet.cell(row=i + 2, column=3).value
        if human_answer is not None:
            num_label += 1
            if human_answer == answer["answer"]:
                machine_cell.fill = PatternFill(start_color=GREEN, fill_type="solid")
                num_cor += 1
            else:
                machine_cell.fill = PatternFill(start_color=RED, fill_type="solid")
        worksheet.cell(row=i + 2, column=5).value = answer["prob"]
        worksheet.cell(row=i + 2, column=6).value = answer["start"]
        worksheet.cell(row=i + 2, column=7).value = answer["end"]
        worksheet.cell(row=i + 2, column=8).value = answer["version"]
    # workbook.save(DATA_PATH + "/qa_sample_" + str(file_index).zfill(4) + "_update.xlsx")
    workbook.save(full_path)
    if num_label > 0:
        acc = num_cor / num_label * 100.0
    else:
        acc = -1
    return {"File": file_name, "#Records": num_qa, "%Label": 100.0 * num_label / num_qa,
            "%Acc": acc, "#Label": num_label, "#Correct": num_cor}


squad = load_dataset("squad")
print(squad["train"])
print(len(squad["train"]))
n = len(squad["train"])
batch_size = 50
start = 0

label_files = range(10)
file_index = 0
from transformers import pipeline

# model_dir = "./temp"
model_dir = "./srcocotero/tiny-bert-qa_run00"
nlp = pipeline('question-answering', model=model_dir, tokenizer=model_dir)
total_label, total_cor = 0, 0
report_data = []
while start < n:
    end = min(n, start + batch_size)
    if file_index in label_files:
        inputs = []
        for i in range(start, end):
            inputs.append({'question': squad["train"][i]["question"],
                           'context': squad["train"][i]["context"]})
        predicts = nlp.predict(inputs)
        for i, input in enumerate(inputs):
            input["answer"] = predicts[i]['answer']
            input["prob"] = predicts[i]['score']
            input["start"] = predicts[i]['start']
            input["end"] = predicts[i]['end']
            input["version"] = "V1.0"
        stat_report = write_label(inputs, file_index)
        report_data.append(stat_report)
        total_label += stat_report["#Label"]
        total_cor += stat_report["#Correct"]
    start = end
    file_index += 1
print("Total acc:", total_cor / total_label)
from utils import acc_report

acc_report(DATA_PATH + "/acc_report", report_data)
