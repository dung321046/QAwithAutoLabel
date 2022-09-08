import random

import openpyxl
from datasets import load_dataset
from openpyxl.styles import Alignment

DATA_PATH = "./data"
LOG_FILE = DATA_PATH + "/log.txt"


def write_label(answers, file_index, per=1.0):
    file_name = DATA_PATH + "/qa_sample_" + str(file_index).zfill(4) + ".xlsx"

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook["Sheet1"]
    for i, answer in enumerate(answers):
        if per < random.uniform(0, 1):
            worksheet.cell(row=i + 2, column=3).value = answer["answer"]
            worksheet.cell(row=i + 2, column=3).alignment = Alignment(wrap_text=True, vertical='top')
    # workbook.save(DATA_PATH + "/qa_sample_" + str(file_index).zfill(4) + "_upate.xlsx")
    workbook.save(file_name)


squad = load_dataset("squad")
print(squad["train"])
print(len(squad["train"]))
n = len(squad["train"])
batch_size = 50
start = 0

label_files = [0, 1, 2]
partial_label_files = [4, 5, 9]
file_index = 0

while start < n:
    end = min(n, start + batch_size)
    if file_index in label_files:
        inputs = []
        for i in range(start, end):
            inputs.append({'question': squad["train"][i]["question"],
                           'context': squad["train"][i]["context"],
                           'answer': squad["train"][i]["answers"]['text'][0]})
        write_label(inputs, file_index)
    elif file_index in partial_label_files:
        inputs = []
        for i in range(start, end):
            inputs.append({'question': squad["train"][i]["question"],
                           'context': squad["train"][i]["context"],
                           'answer': squad["train"][i]["answers"]['text'][0]})
        write_label(inputs, file_index, per=0.5)
    start = end
    file_index += 1
