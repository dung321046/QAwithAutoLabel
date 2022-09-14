import numpy as np
import torch
from transformers import AutoModelForQuestionAnswering, TrainingArguments
from qa_trainer import QATrainer
import wandb

torch.cuda.empty_cache()
wandb.login(key="17cc4e6449d9bc429d81a90211f21adf938bd629")
wandb.init()
run_name = wandb.run.name
model_name = "srcocotero/tiny-bert-qa"
model = AutoModelForQuestionAnswering.from_pretrained(model_name)

from transformers import AutoTokenizer

tokenizer = AutoTokenizer.from_pretrained(model_name)
from transformers import DefaultDataCollator

data_collator = DefaultDataCollator()


def preprocess_function(examples):
    questions = [q.strip() for q in examples["question"]]
    inputs = tokenizer(
        questions,
        examples["context"],
        max_length=384,
        truncation="only_second",
        return_offsets_mapping=True,
        padding="max_length",
    )
    inputs["id"] = examples["id"]
    inputs["context"] = examples["context"]
    inputs["questions"] = questions
    offset_mapping = inputs.pop("offset_mapping")
    answers = examples["answers"]
    start_positions = []
    end_positions = []

    for i, offset in enumerate(offset_mapping):
        answer = answers[i]
        start_char = answer["answer_start"][0]
        end_char = answer["answer_start"][0] + len(answer["text"][0])
        sequence_ids = inputs.sequence_ids(i)

        # Find the start and end of the context
        idx = 0
        while sequence_ids[idx] != 1:
            idx += 1
        context_start = idx
        while sequence_ids[idx] == 1:
            idx += 1
        context_end = idx - 1

        # If the answer is not fully inside the context, label it (0, 0)
        if offset[context_start][0] > end_char or offset[context_end][1] < start_char:
            start_positions.append(0)
            end_positions.append(0)
        else:
            # Otherwise it's the start and end token positions
            idx = context_start
            while idx <= context_end and offset[idx][0] <= start_char:
                idx += 1
            start_positions.append(idx - 1)

            idx = context_end
            while idx >= context_start and offset[idx][1] >= end_char:
                idx -= 1
            end_positions.append(idx + 1)

    inputs["start_positions"] = start_positions
    inputs["end_positions"] = end_positions
    return inputs


DATA_PATH = "./data"

import openpyxl


def read_labeled_idx(file_index, start_idx):
    file_name = DATA_PATH + "/qa_sample_" + str(file_index).zfill(4) + ".xlsx"
    idx = []
    n = 0
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook["Sheet1"]
    for i in range(100):
        answer = worksheet.cell(row=i + 2, column=3).value
        context = worksheet.cell(row=i + 2, column=1).value
        if context is None:
            break
        n += 1
        if answer:
            idx.append(start_idx + i)
    return idx, n


import pickle

with open("tokenized_squad.pk", "rb") as fr:
    tokenized_squad = pickle.load(fr)
print(type(tokenized_squad))
print(tokenized_squad)
training_args = TrainingArguments(
    output_dir="./results",
    evaluation_strategy="epoch",
    learning_rate=2e-5,
    per_device_train_batch_size=16,
    per_device_eval_batch_size=16,
    num_train_epochs=1000,
    weight_decay=0.01,
)
subtrain_idx = []
start_idx = 0
for i in range(20):
    idx, nrecord = read_labeled_idx(i, start_idx)
    subtrain_idx.extend(idx)
    start_idx += nrecord
subtrain = tokenized_squad["train"].select(subtrain_idx)
subtest = tokenized_squad["validation"].select(range(100))

trainer = QATrainer(
    model=model,
    args=training_args,
    train_dataset=subtrain,
    eval_dataset=subtest,
    tokenizer=tokenizer,
    data_collator=data_collator,
)

trainer.train()
predict = trainer.predict(subtest)
start_pred = np.argmax(predict.predictions[0], axis=1)
end_pred = np.argmax(predict.predictions[1], axis=1)
n = len(start_pred)
start_acc, end_acc = np.zeros(n), np.zeros(n)
for i in range(n):
    if start_pred[i] == predict.label_ids[0][i]:
        start_acc[i] = 1
    if end_pred[i] == predict.label_ids[1][i]:
        end_acc[i] = 1
print("Acc start", sum(start_acc) / n)
print("Acc end", sum(end_acc) / n)
accs = [end_acc[i] and start_acc[i] for i in range(n)]
wandb.log({"acc": sum(accs) / n})
print("Acc:", sum(accs) / n)
trainer.save_model(model_name + "_active00")
